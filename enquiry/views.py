# views.py
from io import BytesIO

import pytz
from accounts.decorators import manager_required, counselor_required

import calendar
import csv
import decimal
from datetime import datetime, date, timedelta

from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db.models import Sum, Count, Q, F
from django.http import HttpResponse, JsonResponse, HttpResponseBadRequest
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import ensure_csrf_cookie
from django.contrib import messages
from django.urls import reverse_lazy
from django.views.generic import ListView

from pytz import timezone as pytz_timezone  # for IST

from .models import Enquiry, Comment, Counsellor
from .forms import EnquiryForm, CounsellorForm, BulkEnquiryForm
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin

IST = pytz_timezone('Asia/Kolkata')


def get_ist_time():
    """Return current datetime in IST"""
    return datetime.now(IST)

@login_required(login_url='accounts:login')
def home(request):
    now_ist = get_ist_time()
    today = now_ist.date()
    pending_count = Enquiry.objects.filter(status='pending').count()
    joined_count = Enquiry.objects.filter(status='joined').count()
    dropout_count = Enquiry.objects.filter(status='dropout').count()
    today_followup_count = Enquiry.objects.filter(followup_date=today, status='pending').count()
    recent_enquiries = Enquiry.objects.all().order_by('-created_at')[:5]
    next_week = today + timedelta(days=7)
    upcoming_followups = Enquiry.objects.filter(
        followup_date__gte=today,
        followup_date__lte=next_week,
        status='pending'
    ).order_by('followup_date')[:5]
    total_count = max(pending_count + joined_count + dropout_count, 1)
    pending_percentage = int((pending_count / total_count) * 200)
    joined_percentage = int((joined_count / total_count) * 200)
    dropout_percentage = int((dropout_count / total_count) * 200)
    context = {
        'pending_count': pending_count,
        'joined_count': joined_count,
        'dropout_count': dropout_count,
        'today_followup_count': today_followup_count,
        'recent_enquiries': recent_enquiries,
        'upcoming_followups': upcoming_followups,
        'pending_percentage': pending_percentage,
        'joined_percentage': joined_percentage,
        'dropout_percentage': dropout_percentage,
    }
    return render(request, 'home.html', context)

@login_required(login_url='accounts:login')
def enquiry_success(request):
    return HttpResponse('Enquiry added successfully!')

@login_required(login_url='accounts:login')
def counsellor_success(request):
    return HttpResponse('Counsellor added successfully!')

@login_required(login_url='accounts:login')
def add_enquiry(request):
    """
    View to add a new enquiry
    """
    if request.method == 'POST':
        form = EnquiryForm(request.POST)
        bypass_mobile_check = request.POST.get('bypass_mobile_check') == 'true'
        
        if form.is_valid():
            # Check if mobile number already exists
            mobile = form.cleaned_data.get('mobile')
            if mobile and not bypass_mobile_check:
                mobile_clean = ''.join(filter(str.isdigit, mobile))[-10:]
                existing = Enquiry.objects.filter(mobile__endswith=mobile_clean).first()
                if existing:
                    messages.warning(request, f"A record with this mobile number already exists for {existing.name}. Please confirm if you want to continue.")
                    return render(request, 'add_enquiry.html', {
                        'form': form,
                        'mobile_exists': True,
                        'existing_student': existing.name
                    })
            
            # Save the form
            enquiry = form.save()
            messages.success(request, f"Enquiry for {enquiry.name} has been added successfully!")
            return redirect('enquiry:enquary_list')  # Redirect to list view
    else:
        form = EnquiryForm()
    
    return render(request, 'add_enquiry.html', {
        'form': form,
    })

@manager_required
def add_counsellor(request):
    if request.method == 'POST':
        form = CounsellorForm(request.POST)
        if form.is_valid():
            counsellor = form.save(commit=False)
            counsellor.user = request.user  # or assign another user
            counsellor.save()
            return redirect('enquiry:counsellor_success')
    else:
        form = CounsellorForm()
    return render(request, 'add_counsellor.html', {'form': form})

@login_required(login_url='accounts:login')
def today_followups(request):
    today = get_ist_time().date()
    enquiries = Enquiry.objects.filter(followup_date=today, status='pending')
    return render(request, 'today_followups.html', {'enquiries': enquiries, 'today': today})



@login_required(login_url='accounts:login')
def joined_students(request):
    students = Enquiry.objects.filter(status='joined').order_by('-created_at')
    return render(request, 'joined_students.html', {'enquiries': students})

@login_required(login_url='accounts:login')
def dropout_students(request):
    students = Enquiry.objects.filter(status='dropout').order_by('-created_at')
    return render(request, 'dropout_students.html', {'enquiries': students})

@login_required(login_url='accounts:login')
def pending_enquiries(request):
    enquiries = Enquiry.objects.filter(status='pending').order_by('-followup_date')
    return render(request, 'pending_enquiries.html', {'enquiries': enquiries})


from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from datetime import timedelta
import calendar

from .models import Enquiry



@login_required(login_url='accounts:login')
def dashboard(request):
    now_ist = get_ist_time()
    today = now_ist.date()

    # Debug logs
    print("IST Now:", now_ist)
    print("Today (IST Date):", today)

    current_month = today.month
    current_year = today.year
    current_day = today.day

    # Status counts
    pending_count = Enquiry.objects.filter(status='pending').count()
    joined_count = Enquiry.objects.filter(status='joined').count()
    dropout_count = Enquiry.objects.filter(status='dropout').count()
    total_enquiries = pending_count + joined_count + dropout_count

    # Chart data
    max_height = 200
    pending_percentage = (pending_count / total_enquiries * max_height) if total_enquiries else 0
    joined_percentage = (joined_count / total_enquiries * max_height) if total_enquiries else 0
    dropout_percentage = (dropout_count / total_enquiries * max_height) if total_enquiries else 0

    # Today's followups - correct for DateField
    today_followup_qs = Enquiry.objects.filter(followup_date=today)
    today_followup_count = today_followup_qs.count()
    print("Today's followups (QS):", today_followup_qs)
    print("Count:", today_followup_count)

    # Recent enquiries
    recent_enquiries = Enquiry.objects.order_by('-created_at')[:10]

    # Upcoming followups
    next_week = today + timedelta(days=7)
    upcoming_followups = Enquiry.objects.filter(
        followup_date__gte=today,
        followup_date__lte=next_week,
        status='pending'
    ).order_by('followup_date')[:10]

    # Calendar view
    cal = calendar.monthcalendar(current_year, current_month)
    month_name = calendar.month_name[current_month]
    weekday_names = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']

    followup_dates = list(
        Enquiry.objects.filter(
            followup_date__year=current_year,
            followup_date__month=current_month
        ).values_list('followup_date__day', flat=True).distinct()
    )

    context = {
        'pending_count': pending_count,
        'joined_count': joined_count,
        'dropout_count': dropout_count,
        'total_enquiries': total_enquiries,
        'today_followup_count': today_followup_count,
        'pending_percentage': pending_percentage,
        'joined_percentage': joined_percentage,
        'dropout_percentage': dropout_percentage,
        'recent_enquiries': recent_enquiries,
        'upcoming_followups': upcoming_followups,
        'calendar_data': cal,
        'month_name': month_name,
        'current_year': current_year,
        'current_month': current_month,
        'current_day': current_day,
        'followup_dates': followup_dates,
        'weekday_names': weekday_names,
    }

    return render(request, 'dashboard.html', context)


@login_required(login_url='accounts:login')
def get_calendar_data(request):
    now_ist = get_ist_time()
    month = int(request.GET.get('month', now_ist.month))
    year = int(request.GET.get('year', now_ist.year))
    if month < 1:
        month = 12
        year -= 1
    elif month > 12:
        month = 1
        year += 1
    cal = calendar.monthcalendar(year, month)
    month_name = calendar.month_name[month]
    current_day = now_ist.day if (now_ist.month == month and now_ist.year == year) else None
    followup_dates = Enquiry.objects.filter(
        followup_date__year=year,
        followup_date__month=month
    ).values_list('followup_date__day', flat=True).distinct()
    calendar_data = []
    for week in cal:
        week_data = []
        for day in week:
            if day == 0:
                week_data.append({'day': '', 'is_today': False, 'has_events': False})
            else:
                week_data.append({
                    'day': day,
                    'is_today': day == current_day,
                    'has_events': day in followup_dates
                })
        calendar_data.append(week_data)
    today = now_ist.date()
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month + 1, 1) - timedelta(days=1)
    month_followups = Enquiry.objects.filter(
        followup_date__gte=max(start_date, today),
        followup_date__lte=end_date,
        status='pending'
    ).order_by('followup_date')[:10]
    followups_data = []
    for followup in month_followups:
        followups_data.append({
            'name': followup.name,
            'course': followup.course.name if hasattr(followup.course, 'name') else str(followup.course),
            'date': followup.followup_date.strftime('%Y-%m-%d'),
            'id': followup.id
        })
    response_data = {
        'calendar': calendar_data,
        'month_name': month_name,
        'year': year,
        'month': month,
        'followups': followups_data
    }
    return JsonResponse(response_data)

@login_required(login_url='accounts:login')
def enquiry_details(request, enquiry_id):
    enquiry = get_object_or_404(Enquiry, id=enquiry_id)
    comments = enquiry.comments.all()
    return render(request, 'enquiry_details.html', {'enquiry': enquiry, 'comments': comments})


class EnquiryListView(LoginRequiredMixin, ListView):
    model = Enquiry
    template_name = 'enquiry_list.html'
    context_object_name = 'enquiries'
    login_url = 'accounts:login'

    def get_queryset(self):
        return Enquiry.objects.select_related('counsellor').prefetch_related('comments').all()

from django.http import HttpResponse
import openpyxl
from openpyxl.styles import NamedStyle, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from django.contrib.auth.decorators import login_required

@login_required(login_url='accounts:login')
def download_dummy_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enquiry Template"

    # Headers matching the upload expectations
    headers = [
        'Name',
        'Mobile',
        'Subject (e.g. java_full_stack, python_full_stack)',
        'Status (e.g. joined, pending, dropout)',
        'Enquiry Type (e.g. direct_telephonic, someone_walkin)',
        'Counsellor Name',
        'Followup Date (DD-MM-YYYY)',
        'Enquiry Date (DD-MM-YYYY)',
        'Parent Number',
        'Target Fees',
        'Fees Paid',
        'Due Date (DD-MM-YYYY)',
    ]
    ws.append(headers)

    # Bold header row
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Today's date for consistent formatting
    today = datetime.now().strftime('%d-%m-%Y')

    # Add multiple example rows
    example_rows = [
        ['John Doe', '9876543210', 'java_full_stack', 'pending', 'direct_telephonic', 'Jane Smith', today, today, '9123456789', '45000.00', '10000.00', today],
        ['Alice Brown', '9876500001', 'python_full_stack', 'joined', 'someone_walkin', 'Robert Wilson', today, today, '9876500012', '40000.00', '40000.00', today],
        ['Mark Taylor', '9876500002', 'ui_ux_design', 'dropout', 'direct_telephonic', 'Emily Davis', today, today, '9876500023', '30000.00', '5000.00', today],
        ['Sara Khan', '9876500003', 'data_science', 'pending', 'whatsapp', 'Mohit Reddy', today, today, '9876500034', '60000.00', '0.00', ''],
        ['Amit Verma', '9876500004', 'mern_stack', 'joined', 'someone_walkin', 'Jane Smith', today, today, '9876500045', '50000.00', '25000.00', today],
        ['Priya Patel', '9876500005', 'testing', 'joined', 'email_referral', 'Robert Wilson', today, today, '9876500056', '35000.00', '35000.00', today],
        ['Rohan Iyer', '9876500006', 'python_full_stack', 'pending', 'walkin', 'Emily Davis', today, today, '9876500067', '40000.00', '0.00', ''],
    ]

    for row in example_rows:
        ws.append(row)

    # Date formatting style
    date_style = NamedStyle(name="datetime_style", number_format='DD-MM-YYYY')
    date_columns = [7, 8, 12]  # Followup Date, Enquiry Date, Due Date

    for col_idx in date_columns:
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.style = date_style

    # Auto-adjust column widths
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        adjusted_width = max_length + 5
        ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    # Return Excel file as response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="enquiry_upload_template.xlsx"'
    wb.save(response)
    return response


from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from datetime import datetime
import openpyxl
import pytz

from .forms import BulkEnquiryForm
from .models import Enquiry, Counsellor


def parse_date(value):
    indian_timezone = pytz.timezone('Asia/Kolkata')
    if isinstance(value, datetime):
        return value.astimezone(indian_timezone).date()
    elif value:
        try:
            parsed = datetime.strptime(str(value), "%d-%m-%Y")
            localized = indian_timezone.localize(parsed)
            return localized.date()
        except ValueError:
            raise ValueError(f"Invalid date format: {value}")
    return None


def normalize_choice(value, valid_choices):
    value_cleaned = str(value).strip().lower().replace(' ', '_').replace('-', '_')
    for key in valid_choices:
        if key == value_cleaned:
            return valid_choices[key]
    raise ValueError(f"Invalid value: '{value}'")


def bulk_enquiry_upload(request):
    if request.method == 'POST':
        form = BulkEnquiryForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                # Current IST timestamp
                ist_now = timezone.now().astimezone(pytz.timezone('Asia/Kolkata'))
                print("Upload initiated at IST:", ist_now.strftime("%Y-%m-%d %H:%M:%S"))

                wb = openpyxl.load_workbook(request.FILES['excel_file'])
                ws = wb.active

                subject_map = {k.lower().replace(' ', '_').replace('-', '_'): k for k, _ in Enquiry.SUBJECT_CHOICES}
                status_map = {k.lower().replace(' ', '_').replace('-', '_'): k for k, _ in Enquiry.STATUS_CHOICES}
                enquiry_type_map = {k.lower().replace(' ', '_').replace('-', '_'): k for k, _ in Enquiry.ENQUIRY_TYPE_CHOICES}

                created_count = 0
                updated_count = 0
                error_rows = []
                rows_to_process = []

                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        if not any(row):
                            continue

                        name = str(row[0]).strip() if row[0] else ''
                        mobile = str(row[1]).strip() if row[1] else ''
                        if not mobile:
                            raise ValueError("Mobile number is required")

                        clean_mobile = ''.join(filter(str.isdigit, mobile))
                        if len(clean_mobile) != 10:
                            raise ValueError("Mobile must be a 10-digit number")

                        subject_key = str(row[2]).strip() if row[2] else ''
                        status_key = str(row[3]).strip() if row[3] else 'pending'
                        enquiry_type_key = str(row[4]).strip() if row[4] else 'direct_telephonic'
                        counsellor_name = str(row[5]).strip() if row[5] else ''

                        if not counsellor_name:
                            raise ValueError("Counsellor name is required")

                        followup_date_raw = row[6]
                        enquiry_date_raw = row[7] if len(row) > 7 else None
                        parent_number = str(row[8]).strip() if len(row) > 8 and row[8] else ''
                        target_fees = float(row[9]) if len(row) > 9 and row[9] else None
                        fees_paid = float(row[10]) if len(row) > 10 and row[10] else 0.0
                        due_date_raw = row[11] if len(row) > 11 else None

                        subject = normalize_choice(subject_key, subject_map)
                        status = normalize_choice(status_key, status_map)
                        enquiry_type = normalize_choice(enquiry_type_key, enquiry_type_map)
                        clean_parent_number = ''.join(filter(str.isdigit, parent_number)) if parent_number else ''

                        if status == 'pending':
                            target_fees = 0.0
                            fees_paid = 0.0
                            due_date_raw = None

                        if target_fees is not None and fees_paid > target_fees:
                            raise ValueError("Fees paid cannot exceed target fees.")

                        rows_to_process.append({
                            'row': row,
                            'name': name,
                            'mobile': clean_mobile,
                            'subject': subject,
                            'status': status,
                            'enquiry_type': enquiry_type,
                            'counsellor_name': counsellor_name,
                            'followup_date': followup_date_raw,
                            'enquiry_date': enquiry_date_raw,
                            'parent_number': clean_parent_number,
                            'target_fees': target_fees,
                            'fees_paid': fees_paid,
                            'due_date': due_date_raw,
                        })

                    except Exception as e:
                        error_rows.append({
                            'row_number': row_num,
                            'row_data': row,
                            'error': str(e)
                        })

                if error_rows:
                    return render(request, 'bulk_upload_errors.html', {
                        'form': form,
                        'errors': error_rows
                    })

                # Match counsellors
                counsellor_names = {r['counsellor_name'].strip().lower() for r in rows_to_process}
                query = Q()
                for name in counsellor_names:
                    query |= Q(name__iexact=name)
                counsellors = list(Counsellor.objects.filter(query))
                counsellor_map = {c.name.strip().lower(): c for c in counsellors}

                missing_counsellors = [n for n in counsellor_names if n not in counsellor_map]
                if missing_counsellors:
                    for row in rows_to_process:
                        if row['counsellor_name'].strip().lower() in missing_counsellors:
                            error_rows.append({
                                'row_number': None,
                                'row_data': row['row'],
                                'error': f"Counsellor not found: {row['counsellor_name']}"
                            })
                    return render(request, 'bulk_upload_errors.html', {
                        'form': form,
                        'errors': error_rows
                    })

                mobiles_in_excel = [r['mobile'] for r in rows_to_process]
                existing_records = Enquiry.objects.filter(mobile__in=mobiles_in_excel).values('mobile')
                existing_mobile_set = {e['mobile'] for e in existing_records}

                enquirys_to_create = []
                enquirys_to_update = []

                for row in rows_to_process:
                    mobile_clean = row['mobile']
                    balance = max(row['target_fees'] - row['fees_paid'], 0) if row['target_fees'] is not None else None
                    parsed_updates = {
                        'name': row['name'],
                        'subject': row['subject'],
                        'status': row['status'],
                        'enquiry_type': row['enquiry_type'],
                        'counsellor': counsellor_map[row['counsellor_name'].strip().lower()],
                        'followup_date': parse_date(row['followup_date']),
                        'enquiry_date': parse_date(row['enquiry_date']),
                        'parent_number': row['parent_number'],
                        'target_fees': row['target_fees'],
                        'fees_paid': row['fees_paid'],
                        'fees_balance': balance,
                        'due_date': parse_date(row['due_date']),
                    }

                    if mobile_clean in existing_mobile_set:
                        enquirys_to_update.append({'mobile': mobile_clean, 'updates': parsed_updates})
                        updated_count += 1
                    else:
                        enquirys_to_create.append(Enquiry(mobile=mobile_clean, **parsed_updates))
                        created_count += 1

                with transaction.atomic():
                    if enquirys_to_create:
                        Enquiry.objects.bulk_create(enquirys_to_create)
                    if enquirys_to_update:
                        for item in enquirys_to_update:
                            Enquiry.objects.filter(mobile=item['mobile']).update(**item['updates'])

                messages.success(request, f"✅ Created: {created_count}, ✏️ Updated: {updated_count}")
                return redirect('enquiry:bulk_enquiry_upload')

            except Exception as e:
                messages.error(request, f"❌ Error processing file: {str(e)}")

    else:
        form = BulkEnquiryForm()

    return render(request, 'bulk_upload.html', {'form': form})



from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from .models import Enquiry, Counsellor
from .forms import EnquiryForm

@login_required(login_url='accounts:login')
def edit_enquiry(request, pk):
    enquiry = get_object_or_404(Enquiry, pk=pk)
    
    if request.method == 'POST':
        form = EnquiryForm(request.POST, instance=enquiry)

        if form.is_valid():
            form.save()
            messages.success(request, 'Enquiry updated successfully!')
            return redirect('enquiry:enquary_list')

        else:
            # Clear fees-related errors if status is not 'joined'
            status = form.data.get('status', '')
            if status != 'joined':
                for field in ['target_fees', 'fees_paid', 'due_date']:
                    if field in form.errors:
                        del form.errors[field]

            messages.warning(request, 'Please correct the errors below.')

    else:
        form = EnquiryForm(instance=enquiry)

    context = {
        'form': form,
        'enquiry': enquiry,
        'counsellor': form.fields['counsellor'].queryset,
        'status_choices': Enquiry.STATUS_CHOICES,
        'subject_choices': Enquiry.SUBJECT_CHOICES,
        'enquiry_type_choices': Enquiry.ENQUIRY_TYPE_CHOICES,
    }

    return render(request, 'edit_enquiry.html', context)
# ========== Monthly Student List ==========
@login_required(login_url='accounts:login')
def monthly_student_list(request):
    selected_month = request.GET.get('month', None)
    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('q', '')
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    now = datetime.now()
    current_year = now.year
    current_month = now.month
    years = [2025] if current_year < 2025 else list(range(2025, current_year + 1))
    
    enquiries = Enquiry.objects.select_related('counsellor').order_by('-created_at')
    
    if selected_month:
        year, month = map(int, selected_month.split('-'))
        enquiries = enquiries.filter(created_at__year=year, created_at__month=month)
    
    if status_filter:
        enquiries = enquiries.filter(status=status_filter)
    
    if search_query:
        enquiries = enquiries.filter(
            Q(name__icontains=search_query) |
            Q(mobile__icontains=search_query) |
            Q(subject__icontains=search_query)
        )
    
    paginator = Paginator(enquiries, 10)
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    total_count = page_obj.paginator.count
    joined_count = enquiries.filter(status='joined').count()
    pending_count = enquiries.filter(status='pending').count()
    dropout_count = enquiries.filter(status='dropout').count()
    
    context = {
        'enquiries': page_obj,
        'selected_month': selected_month,
        'status_filter': status_filter,
        'search_query': search_query,
        'total_count': total_count,
        'joined_count': joined_count,
        'pending_count': pending_count,
        'dropout_count': dropout_count,
        'month_choices': [
            (f"{year}-{str(month).zfill(2)}", f"{calendar.month_name[month]} {year}")
            for year in years
            for month in range(1, 13)
            if not (year == current_year and month > current_month)
        ]
    }
    
    if is_ajax:
        return JsonResponse({
            'total_count': total_count,
            'joined_count': joined_count,
            'pending_count': pending_count,
            'dropout_count': dropout_count,
        })
    
    return render(request, 'monthly_student_list.html', context)

# ========== Download Monthly Student List ==========
@login_required(login_url='accounts:login')
def download_monthly_excel(request):
    selected_month = request.GET.get('month')
    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('q', '')
    
    enquiries = Enquiry.objects.select_related('counsellor').order_by('-created_at')
    
    if selected_month:
        year, month = map(int, selected_month.split('-'))
        enquiries = enquiries.filter(created_at__year=year, created_at__month=month)
    
    if status_filter:
        enquiries = enquiries.filter(status=status_filter)
    
    if search_query:
        enquiries = enquiries.filter(
            Q(name__icontains=search_query) |
            Q(mobile__icontains=search_query) |
            Q(subject__icontains=search_query)
        )
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Enquiries"
    
    headers = [
        'Name', 'Mobile', 'Course', 'Status',
        'Enquiry Type', 'Counsellor', 'Follow-up Date', 'Enquiry Date'
    ]
    ws.append(headers)
    
    for enquiry in enquiries:
        ws.append([
            enquiry.name,
            enquiry.mobile,
            enquiry.get_subject_display(),
            enquiry.get_status_display(),
            enquiry.get_enquiry_type_display(),
            enquiry.counsellor.name if enquiry.counsellor else '',
            enquiry.followup_date.strftime('%Y-%m-%d') if enquiry.followup_date else '',
            enquiry.created_at.strftime('%Y-%m-%d %H:%M'),
        ])
    
    current_date = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"monthly_enquiries_{selected_month or 'all'}_{current_date}.xlsx"
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

# ========== Monthly Summary Dashboard ==========
from datetime import datetime
from decimal import Decimal
from django.db.models import Sum
import calendar
from .models import Enquiry

@manager_required
def get_monthly_stats(request, month_filter=None, status_filter=None):
    stats = []
    now = datetime.now()
    years = list(range(2025, now.year + 1))
    
    for year in years:
        start_month = 1
        end_month = 12 if year < now.year else now.month
        for month in range(start_month, end_month + 1):
            start_date = datetime(year, month, 1)
            end_date = datetime(year, month, calendar.monthrange(year, month)[1])
            queryset = Enquiry.objects.filter(created_at__range=(start_date, end_date))

            if status_filter:
                queryset = queryset.filter(status=status_filter)

            joined_qs = queryset.filter(status='joined')

            total_students = queryset.count()
            joined_count = joined_qs.count()
            pending_count = queryset.filter(status='pending').count()
            dropout_count = queryset.filter(status='dropout').count()
            
            total_fees_paid = queryset.aggregate(Sum('fees_paid'))['fees_paid__sum'] or Decimal('0.00')
            total_balance = joined_qs.filter(fees_balance__gt=0).aggregate(Sum('fees_balance'))['fees_balance__sum'] or Decimal('0.00')

            stats.append({
                'month': f"{year}-{month:02d}",
                'label': f"{calendar.month_abbr[month]} {year}",
                'total_students': total_students,
                'joined': joined_count,
                'pending': pending_count,
                'dropout': dropout_count,
                'fees_paid': float(total_fees_paid),
                'fees_balance': float(total_balance),
            })
    
    return stats


@manager_required
def monthly_summary_dashboard(request):
    selected_month = request.GET.get('month', None)
    
    status_filter = request.GET.get('status', '')
    stats = get_monthly_stats(request, selected_month, status_filter)
    
    context = {
        'title': 'Monthly Summary',
        'stats': stats,
        'selected_month': selected_month,
        'status_filter': status_filter,
    }
    return render(request, 'monthly_summary_dashboard.html', context)
# ========== Download Monthly Summary Excel ==========
@login_required(login_url='accounts:login')
def download_monthly_summary_excel(request):
    selected_month = request.GET.get('month')
    status_filter = request.GET.get('status')
    stats = get_monthly_stats(request, selected_month, status_filter)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Summary"
    
    headers = ['Month', 'Total Students', 'Joined', 'Pending', 'Dropout', 'Fees Collected', 'Pending Balance']
    ws.append(headers)
    
    for row in stats:
        ws.append([
            row['label'],
            row['total_students'],
            row['joined'],
            row['pending'],
            row['dropout'],
            row['fees_paid'],
            row['fees_balance'],
        ])
    
    current_date = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"monthly_summary_{selected_month or 'all'}_{current_date}.xlsx"
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

from django.db.models import Q, Sum
from decimal import Decimal

def get_monthly_stats_by_counsellor(request, month_filter=None, status_filter=None):
    stats = []
    counsellors = Counsellor.objects.all().prefetch_related('enquiry_set')
    from django.db.models import Q, Sum
from decimal import Decimal

def get_monthly_stats_by_counsellor(request, month_filter=None, status_filter=None):
    stats = []
    counsellors = Counsellor.objects.all().prefetch_related('enquiry_set')

    # List of all telephonic-related enquiry types
    TELEPHONIC_TYPES = [
        'someone_telephonic',
        'direct_telephonic',
        'telephonic_to_walkin',
    ]


    for counsellor in counsellors:
        queryset = Enquiry.objects.filter(counsellor=counsellor)

        if month_filter:
            try:
                year, month = map(int, month_filter.split('-'))
                queryset = queryset.filter(created_at__year=year, created_at__month=month)
            except ValueError:
                pass

        if status_filter:
            queryset = queryset.filter(status=status_filter)

        # Split by visit_type
        telephonic_q = Q(enquiry_type__in=TELEPHONIC_TYPES)
        walkin_q = ~telephonic_q

        # Total Telephonic / Walkin
        total_telephonic = queryset.filter(telephonic_q).count()
        total_walkin = queryset.filter(walkin_q).count()
        
        print(total_telephonic)
        

        # Joined
        joined_telephonic = queryset.filter(telephonic_q, status='joined').count()
        joined_walkin = queryset.filter(walkin_q, status='joined').count()
        joined_total = joined_telephonic + joined_walkin
        print(joined_telephonic)

        # Pending
        pending_telephonic = queryset.filter(telephonic_q, status='pending').count()
        pending_walkin = queryset.filter(walkin_q, status='pending').count()
        pending_total = pending_telephonic + pending_walkin

        # Dropout
        dropout_telephonic = queryset.filter(telephonic_q, status='dropout').count()
        dropout_walkin = queryset.filter(walkin_q, status='dropout').count()
        dropout_total = dropout_telephonic + dropout_walkin

        # Fees
        fees_data = queryset.aggregate(Sum('fees_paid'), Sum('fees_balance'))
        fees_paid = fees_data['fees_paid__sum'] or Decimal('0.00')
        fees_balance = fees_data['fees_balance__sum'] or Decimal('0.00')

        stats.append({
            'counsellor': counsellor,
            'total_telephonic': total_telephonic,
            'total_walkin': total_walkin,
            'joined_telephonic': joined_telephonic,
            'joined_walkin': joined_walkin,
            'joined': joined_total,  # ✅ Precomputed total
            'pending_telephonic': pending_telephonic,
            'pending_walkin': pending_walkin,
            'pending': pending_total,  # ✅ Precomputed total
            'dropout_telephonic': dropout_telephonic,
            'dropout_walkin': dropout_walkin,
            'dropout': dropout_total,  # ✅ Precomputed total
            'fees_paid': float(fees_paid),
            'fees_balance': float(fees_balance),
        })

    return stats

# views.py
import calendar
from datetime import datetime
from django.shortcuts import render
from django.db.models import Q, Sum
from decimal import Decimal

@login_required
@manager_required
def monthly_summary_by_counsellor(request):
    selected_month = request.GET.get('month', '')
    status_filter = request.GET.get('status', '')
    
    # Get current date context
    now = datetime.now()
    current_year = now.year
    current_month = now.month
    
    # Validate and parse selected_month
    try:
        if selected_month:
            year, month = map(int, selected_month.split('-'))
            selected_month = f"{year}-{month:02d}"
        else:
            selected_month = f"{current_year}-{current_month:02d}"
            year, month = current_year, current_month
    except ValueError:
        selected_month = f"{current_year}-{current_month:02d}"
        year, month = current_year, current_month

    # Get stats with fallback
    try:
        stats = get_monthly_stats_by_counsellor(
            request, 
            selected_month, 
            status_filter
        )
    except Exception as e:
        print(f"Error generating stats: {e}")
        stats = []

    # Generate month choices
    years = list(range(2020, current_year + 1))  # Fixed year range
    month_choices = []
    for year_opt in years:
        start_month = 1
        end_month = 12 if year_opt < current_year else current_month
        
        for m in range(start_month, end_month + 1):
            month_label = f"{calendar.month_abbr[m]} {year_opt}"
            month_value = f"{year_opt}-{str(m).zfill(2)}"
            month_choices.append((month_value, month_label))

    context = {
        'title': 'Monthly Summary by Counsellor',
        'stats': stats,
        'selected_month': selected_month,
        'status_filter': status_filter,
        'month_choices': month_choices,
        'current_year': current_year,
        'current_month': current_month,
    }
    
    return render(request, 'monthly_summary_by_counsellor.html', context)


from openpyxl import Workbook
# ========== Download Monthly Summary by Counsellor Excel ==========
@login_required(login_url='accounts:login')
def download_monthly_summary_by_counsellor_excel(request):
    selected_month = request.GET.get('month')
    status_filter = request.GET.get('status')
    stats = get_monthly_stats_by_counsellor(request,selected_month, status_filter)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Counsellor Summary"

    headers = ['Counsellor', 'Total Students', 'Joined', 'Pending', 'Dropout', 'Fees Collected', 'Pending Balance']
    ws.append(headers)

    for row in stats:
        ws.append([
            row['counsellor'].name,  # ✅ Now writes the counsellor's name
            row['total_students'],
            row['joined'],
            row['pending'],
            row['dropout'],
            row['fees_paid'],
            row['fees_balance'],
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"counsellor_summary_{selected_month or 'all'}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


from django.shortcuts import get_object_or_404, render
from django.utils.timezone import make_aware
from datetime import datetime, timedelta
from .models import Enquiry, Counsellor
from django.contrib.auth.decorators import login_required

@login_required(login_url='accounts:login')
def counsellor_monthly_details(request, counsellor_id, year, month):
    counsellor = get_object_or_404(Counsellor, id=counsellor_id)
    
    if not (1 <= month <= 12):
        return render(request, 'error.html', {'message': 'Invalid month'})
    
    try:
        start_date = datetime(year=year, month=month, day=1)
        next_month = start_date.replace(day=28) + timedelta(days=4)
        end_date = next_month.replace(day=1) - timedelta(days=1)
    except ValueError:
        return render(request, 'error.html', {'message': 'Invalid date range'})
    
    start_date = make_aware(start_date)
    end_date = make_aware(datetime.combine(end_date, datetime.max.time()))
    
    enquiries = Enquiry.objects.filter(
        counsellor=counsellor,
        created_at__range=(start_date, end_date)
    ).select_related('counsellor')
    
    joined_count = enquiries.filter(status='joined').count()
    pending_count = enquiries.filter(status='pending').count()
    dropout_count = enquiries.filter(status='dropout').count()
    
    context = {
        'counsellor': counsellor,
        'enquiries': enquiries,
        'month': start_date.strftime('%B'),
        'year': year,
        'joined_count': joined_count,
        'pending_count': pending_count,
        'dropout_count': dropout_count,
    }
    return render(request, 'counsellor_monthly_details.html', context)


@login_required(login_url='accounts:login')
def payment_due_list(request, due_type='today'):
    today = get_ist_time().date()
    queryset = Enquiry.objects.filter(
        status='joined'
    ).filter(
        Q(fees_paid__lt=F('target_fees'))
    ).select_related('counsellor').order_by('due_date')
    if due_type == 'today':
        queryset = queryset.filter(due_date=today)
    elif due_type == 'past':
        queryset = queryset.filter(due_date__lt=today)
    search_counsellor = request.GET.get('search_counsellor', '').strip()
    if search_counsellor:
        queryset = queryset.filter(counsellor__name__icontains=search_counsellor)
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    context = {
        'title': 'Payment Dues',
        'page_obj': page_obj,
        'search_counsellor': search_counsellor,
        'due_type': due_type,
        'today': today,
    }
    return render(request, 'payment_due_list.html', context)

@login_required(login_url='accounts:login')
def export_payment_due_csv(request, due_type='today'):
    today = timezone.now().date()
    if due_type == 'today':
        date_filter = {'due_date': today}
        filename = f"today_payment_due_{today}.csv"
    elif due_type == 'past':
        date_filter = {'due_date__lt': today}
        filename = f"past_payment_due_{today}.csv"
    else:
        date_filter = {}
        filename = f"all_payment_due_{today}.csv"
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow([
        'Student Name', 'Counsellor', 'Course', 'Total Fees',
        'Fees Paid', 'Balance', 'Due Date'
    ])
    queryset = Enquiry.objects.filter(
        status='joined',
        fees_balance__gt=0,
        **date_filter
    ).select_related('counsellor')
    for student in queryset:
        writer.writerow([
            student.name,
            student.counsellor.name,
            student.get_subject_display(),
            student.target_fees,
            student.fees_paid,
            student.fees_balance,
            student.due_date
        ])
    return response


@login_required(login_url='accounts:login')
def update_due_date(request):
    if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'success': False,
            'message': 'Invalid request method'
        }, status=400)
    try:
        student_id = request.POST.get('student_id')
        new_due_date_str = request.POST.get('new_due_date')
        if not student_id or not new_due_date_str:
            return JsonResponse({
                'success': False,
                'message': 'Missing required fields'
            }, status=400)
        try:
            new_due_date = datetime.strptime(new_due_date_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({
                'success': False,
                'message': 'Invalid date format'
            }, status=400)
        today = datetime.today().date()
        if new_due_date < today:
            return JsonResponse({
                'success': False,
                'message': 'Due date cannot be in the past'
            }, status=400)
        enquiry = get_object_or_404(Enquiry, id=student_id)
        enquiry.due_date = new_due_date
        enquiry.save(update_fields=['due_date'])
        return JsonResponse({
            'success': True,
            'message': f'Due date for {enquiry.name} updated successfully',
            'new_due_date': new_due_date.strftime('%b %d, %Y')
        })
    except Exception as e:
        print(f"Error updating due date: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': 'An error occurred while processing your request'
        }, status=500)
        
        
from datetime import datetime
import decimal

@login_required(login_url='accounts:login')
def update_payment(request):
    try:
        enquiry_id = request.POST.get('enquiry_id')
        payment_amount = request.POST.get('payment_amount')
        target_fees = request.POST.get('target_fees')
        due_date_str = request.POST.get('due_date')  # <-- New line

        if not enquiry_id or not payment_amount:
            return JsonResponse({'success': False, 'error': 'Missing required fields.'})

        try:
            amount = decimal.Decimal(payment_amount)
            target = decimal.Decimal(target_fees) if target_fees else None
        except decimal.InvalidOperation:
            return JsonResponse({'success': False, 'error': 'Invalid payment amount or target fees.'})

        if amount <= 0:
            return JsonResponse({'success': False, 'error': 'Payment must be greater than zero.'})

        enquiry = Enquiry.objects.get(id=enquiry_id)

        if target is not None:
            enquiry.target_fees = target

        if enquiry.target_fees is not None and (enquiry.fees_paid + amount) > enquiry.target_fees:
            return JsonResponse({
                'success': False,
                'error': f'Payment exceeds target fees of ₹{enquiry.target_fees}.'
            })

        # Process due_date
        if due_date_str:
            try:
                new_due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
                if new_due_date < datetime.today().date():
                    return JsonResponse({'success': False, 'error': 'Due date cannot be in the past.'})
                enquiry.due_date = new_due_date
            except ValueError:
                return JsonResponse({'success': False, 'error': 'Invalid due date format.'})

        enquiry.fees_paid += amount
        enquiry.fees_paid_date = datetime.today().date()

        if enquiry.target_fees is not None:
            enquiry.fees_balance = enquiry.target_fees - enquiry.fees_paid
        else:
            enquiry.fees_balance = None

        enquiry.save(update_fields=['target_fees', 'fees_paid', 'fees_balance', 'fees_paid_date', 'due_date'])

        return JsonResponse({
            'success': True,
            'fees_paid': float(enquiry.fees_paid),
            'fees_balance': float(enquiry.fees_balance),
            'message': 'Payment and due date updated successfully.'
        })

    except Enquiry.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Enquiry not found.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def get_fees_details(request, enquiry_id):
    try:
        enquiry = Enquiry.objects.get(id=enquiry_id)
        return JsonResponse({
            'fees_paid': str(enquiry.fees_paid),
            'target_fees': str(enquiry.target_fees) if enquiry.target_fees else None,
            'fees_balance': str(enquiry.fees_balance) if enquiry.fees_balance else None
        })
    except Enquiry.DoesNotExist:
        return JsonResponse({'error': 'Enquiry not found'}, status=404)

@require_POST
@login_required(login_url='accounts:login')
def add_comment(request):
    try:
        enquiry_id = request.POST.get('enquiry_id')
        comment_text = request.POST.get('comment')
        if not enquiry_id or not comment_text:
            return JsonResponse({
                'status': 'error',
                'message': 'Missing required parameters'
            }, status=400)
        enquiry = Enquiry.objects.get(id=enquiry_id)
        comment = Comment.objects.create(
            enquiry=enquiry,
            comment=comment_text,
        )
        return JsonResponse({
            'status': 'success',
            'truncated_comment': comment.comment[:40] + ('...' if len(comment.comment) > 40 else '')
        })
    except Enquiry.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Enquiry not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)

@login_required(login_url='accounts:login')
def update_followup_date(request):
    if request.method == 'POST':
        enquiry_id = request.POST.get('enquiry_id')
        new_date = request.POST.get('followup_date')
        try:
            enquiry = Enquiry.objects.get(id=enquiry_id)
            enquiry.followup_date = new_date
            enquiry.save()
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'})

@login_required(login_url='accounts:login')
def update_comment(request):
    if request.method == "POST":
        enquiry_id = request.POST.get("enquiry_id")
        new_comment = request.POST.get("comment")
        try:
            enquiry = Enquiry.objects.get(id=enquiry_id)
            enquiry.comment = new_comment
            enquiry.save()
            return JsonResponse({'status': 'success'})
        except Enquiry.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Enquiry not found.'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'})

@login_required(login_url='accounts:login')
def get_comments(request):
    enquiry_id = request.GET.get('enquiry_id')
    enquiry = Enquiry.objects.get(id=enquiry_id)
    comments = enquiry.comments.all().order_by('-created_at')
    return render(request, 'comments_partial.html', {'comments': comments})


@login_required(login_url='accounts:login')
def update_status(request):
    if request.method == 'POST':
        enquiry_id = request.POST.get('enquiry_id')
        new_status = request.POST.get('new_status')

        try:
            enquiry = Enquiry.objects.get(id=enquiry_id)
            old_status = enquiry.status  # capture current status before change

            enquiry.status = new_status

            # If status changed to 'joined' and it wasn't previously joined
            if new_status == 'joined' and old_status != 'joined' and not enquiry.joined_date:
                ist = pytz.timezone('Asia/Kolkata')
                enquiry.joined_date = timezone.now().astimezone(ist).date()

            enquiry.save()

            return JsonResponse({
                'status': 'success',
                'message': 'Status updated successfully',
                'new_status_display': enquiry.get_status_display(),
                'new_status': new_status
            })

        except Enquiry.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Enquiry not found'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)

from django.http import JsonResponse
from datetime import datetime
from .models import Enquiry
@login_required(login_url='accounts:login')
def get_followups(request):
    date_str = request.GET.get('date', '')
    
    try:
        if len(date_str.split('-')) == 3:
            # Format: YYYY-MM-DD
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
            followups = Enquiry.objects.filter(followup_date=date_obj)
        else:
            # Format: YYYY-MM
            year, month = map(int, date_str.split('-'))
            followups = Enquiry.objects.filter(followup_date__year=year, followup_date__month=month)
        
        followup_data = []
        for followup in followups:
            followup_data.append({
                'id': followup.id,
                'name': followup.name,
                'course': followup.subject,
                'phone': followup.mobile,
                'followup_date': followup.followup_date.strftime('%Y-%m-%d')
            })
        
        return JsonResponse({
            'status': 'success',
            'followups': followup_data
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })

@login_required(login_url='accounts:login')
def get_calendar_data(request):
    month = int(request.GET.get('month', datetime.now().month))
    year = int(request.GET.get('year', datetime.now().year))

    # Determine number of days in the month
    if month == 2:
        is_leap = year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)
        days_in_month = 29 if is_leap else 28
    elif month in [4, 6, 9, 11]:
        days_in_month = 30
    else:
        days_in_month = 31

    # Calculate first weekday (Sunday = 0)
    first_day = datetime(year, month, 1).weekday()
    first_day = (first_day + 1) % 7

    # Get all distinct follow-up days
    followups = Enquiry.objects.filter(
        followup_date__year=year,
        followup_date__month=month
    ).values_list('followup_date__day', flat=True).distinct()
    
    return JsonResponse({
        'days_in_month': days_in_month,
        'first_day_of_month': first_day,
        'followup_dates': list(followups)
    })




from django.http import JsonResponse
from .models import Enquiry
@login_required(login_url='accounts:login')
def check_mobile_exists(request):
    """
    AJAX view to check if a mobile number already exists in the database
    """
    mobile = request.GET.get('mobile', '')
    enquiry_id = request.GET.get('enquiry_id', None)
    
    # Strip any non-digit characters and get the last 10 digits
    mobile = ''.join(filter(str.isdigit, mobile))[-10:]
    
    # Check if the mobile number exists, excluding the current enquiry if we're editing
    if enquiry_id:
        exists = Enquiry.objects.filter(mobile__endswith=mobile).exclude(id=enquiry_id).exists()
    else:
        exists = Enquiry.objects.filter(mobile__endswith=mobile).exists()
    
    if exists:
        # Get the name of the student with this number
        student = Enquiry.objects.filter(mobile__endswith=mobile).first()
        return JsonResponse({
            'exists': True,
            'message': f'This number is already registered for {student.name}',
            'student_name': student.name
        })
    else:
        return JsonResponse({'exists': False})
    
from django.shortcuts import render
from django.db.models import Sum, F
from .models import Enquiry, Counsellor, PaymentHistory
from datetime import date
from django.contrib import messages
# if not already defined, adjust as needed

@manager_required
def daywise_counsellor_summary(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    try:
        start_date = date.fromisoformat(start_date) if start_date else date.today()
        end_date = date.fromisoformat(end_date) if end_date else date.today()
    except ValueError:
        start_date = end_date = date.today()
        messages.error(request, "Invalid date format. Showing today's data.")

    if start_date > end_date:
        messages.error(request, "Start date must be before end date.")
        start_date, end_date = end_date, start_date  # Swap

    counsellors = Counsellor.objects.all().order_by('name')
    summary = []

    for counsellor in counsellors:
        base_queryset = Enquiry.objects.filter(
            counsellor=counsellor,
            created_at__date__range=(start_date, end_date)
        )

        telephonic_all = base_queryset.filter(
            enquiry_type__in=['direct_telephonic', 'someone_telephonic','telephonic_to_walkin']
        )
        walkin_all = base_queryset.filter(
            enquiry_type__in=['direct_walkin', 'someone_walkin' ]
        )

        prev_converted_qs = Enquiry.objects.filter(
            counsellor=counsellor,
            enquiry_date__lt=start_date,
            joined_date__range=(start_date, end_date),
            status='joined'
        )

        total_target_fees = (base_queryset | prev_converted_qs).aggregate(
            total_target=Sum('target_fees')
        )['total_target'] or 0

        enquiry_ids = list((base_queryset | prev_converted_qs).values_list('id', flat=True))
        total_paid = PaymentHistory.objects.filter(
            enquiry_id__in=enquiry_ids,
            payment_date__range=(start_date, end_date)
        ).aggregate(total_paid=Sum('amount_paid'))['total_paid'] or 0

        fees_balance = total_target_fees - total_paid

        summary.append({
            'counsellor_name': counsellor.name,
            'counsellor_mobile': getattr(counsellor, 'mobile', 'N/A'),
            'total_enquiries': base_queryset.count(),
             'counsellor_id': counsellor.id,  # 👈 Add this line
            'telephonic_count': telephonic_all.count(),
            'telephonic_joined': telephonic_all.filter(status='joined').count(),
            'walkin_count': walkin_all.count(),
            'walkin_joined': walkin_all.filter(status='joined').count(),
            'joined_count': base_queryset.filter(status='joined').count(),
            'previous_converted_count': prev_converted_qs.count(),
            'fees_paid': total_paid,
            'fees_target': total_target_fees,
            'fees_balance': fees_balance,
            'pending_count': base_queryset.filter(status='pending').count(),
            'dropout_count': base_queryset.filter(status='dropout').count(),
        })

    # Day-wise collection
    day_wise_fees = PaymentHistory.objects.filter(
        enquiry__status='joined',
        payment_date__range=(start_date, end_date)
    ).values(day=F('payment_date')).annotate(
        total_paid=Sum('amount_paid')
    ).order_by('day')

    total_fees_collected = sum(day['total_paid'] or 0 for day in day_wise_fees)
    total_target_fees = sum(row['fees_target'] for row in summary)
    yet_to_collect = total_target_fees - total_fees_collected

    context = {
        'summary': summary,
        'day_wise_fees': day_wise_fees,
        'start_date': start_date,
        'end_date': end_date,
        'total_fees_collected': total_fees_collected,
        'total_target_fees': total_target_fees,
        'yet_to_collect': yet_to_collect,
    }

    return render(request, 'daywise_summary.html', context)



@login_required(login_url='accounts:login')
def export_to_excel(summary, day_wise_fees, total_collected, total_target, yet_to_collect, start_date, end_date):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Counsellor_Report_{start_date}_to_{end_date}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb = openpyxl.Workbook()
    
    # Counsellor Summary Sheet
    ws1 = wb.active
    ws1.title = "Counsellor Performance"
    ws1.append([
        'Counsellor', 'Mobile', 'Total Enquiries', 'Telephonic', 
        'Telephonic Joined', 'Walk-in', 'Walk-in Joined', 'Total Joined',
        'Pending', 'Dropout', 'Fees Collected', 'Target Fees', 'Balance'
    ])
    
    for row in summary:
        ws1.append([
            row['counsellor_name'],
            row['counsellor_mobile'],
            row['telephonic_count'] + row['walkin_count'],
            row['telephonic_count'],
            row['telephonic_joined'],
            row['walkin_count'],
            row['walkin_joined'],
            row['joined_count'],
            row['pending_count'],
            row['dropout_count'],
            round(row['fees_paid'], 2),
            round(row['fees_target'], 2),
            round(row['fees_balance'], 2),
        ])

    # Financial Summary Sheet
    ws2 = wb.create_sheet("Financial Summary")
    financial_data = [
        ('Total Fees Collected', total_collected),
        ('Total Target Fees', total_target),
        ('Yet to Collect', yet_to_collect),
        ('Collection Efficiency', f"{(total_collected / total_target * 100 if total_target else 0):.1f}%")
    ]
    
    for title, value in financial_data:
        ws2.append([title, round(value, 2)])

    # Day-wise Collection Sheet
    ws3 = wb.create_sheet("Daily Collections")
    ws3.append(['Date', 'Fees Collected'])
    for entry in day_wise_fees:
        ws3.append([entry['created_at__date'], round(entry['total_paid'], 2)])

    wb.save(response)
    return response
from django.utils import timezone

def export_enquiries_view(request):
    import datetime
    if 'start_date' in request.GET and 'end_date' in request.GET:
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')

        # Validate and parse dates
        try:
            start_date = datetime.datetime.strptime(start_date_str, "%d-%m-%Y").date()
            end_date = datetime.datetime.strptime(end_date_str, "%d-%m-%Y").date()
        except ValueError:
            return HttpResponseBadRequest("Invalid date format. Please use DD-MM-YYYY.")

        # Set IST timezone
        ist = pytz.timezone('Asia/Kolkata')

        # Convert to timezone-aware datetimes (start of day and end of day)
        start_datetime = ist.localize(datetime.datetime.combine(start_date, datetime.time.min))
        end_datetime = ist.localize(datetime.datetime.combine(end_date, datetime.time.max))

        # Fetch filtered Enquiries with related data
        enquiries = Enquiry.objects.filter(
            created_at__range=(start_datetime, end_datetime)
        ).select_related(
            'counsellor',
        ).prefetch_related('education_details')

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Enquiries"

        headers = [
            'Name', 'Mobile', 'Subject', 'Status', 'Enquiry Type',
            'Counsellor', 'Created At', 'Followup Date', 'Enquiry Date',
            'Parent Number', 'Native District', 'Target Fees', 'Fees Paid',
            'Fees Balance', 'Due Date', 'Other Subject Name', 'College Name',
            'College Location', 'Visit Type', 'Fees Paid Date',
            'Education Level', 'UG Degree', 'PG Degree', 'Other UG Degree',
            'Other PG Degree', 'Branch', 'Year of Passing', 'Percentage',
        ]
        ws.append(headers)

        for enquiry in enquiries:
            education_info = enquiry.education_details.first() if hasattr(enquiry, 'education_details') else None

            row = [
                enquiry.name,
                enquiry.mobile,
                enquiry.get_subject_display(),
                enquiry.get_status_display(),
                enquiry.get_enquiry_type_display(),
                str(enquiry.counsellor) if enquiry.counsellor else '',
                enquiry.created_at.astimezone(ist).date() if enquiry.created_at else '',
                enquiry.followup_date,
                enquiry.enquiry_date,
                enquiry.parent_number,
                enquiry.native_district_name,
                float(enquiry.target_fees) if enquiry.target_fees else None,
                float(enquiry.fees_paid),
                float(enquiry.fees_balance) if enquiry.fees_balance else None,
                enquiry.due_date,
                enquiry.other_subject_name,
                education_info.college_name if education_info else '',
                education_info.college_place if education_info else '',
                enquiry.get_visit_type_display(),
                enquiry.fees_paid_date,

                # Education Info Fields
                education_info.level if education_info else '',
                education_info.get_ug_degree_display() if education_info and education_info.ug_degree else '',
                education_info.get_pg_degree_display() if education_info and education_info.pg_degree else '',
                education_info.other_ug_degree_name if education_info else '',
                education_info.other_pg_degree_name if education_info else '',
                education_info.branch.upper() if education_info and education_info.branch else '',
                education_info.year_of_passing if education_info else '',
                education_info.percentage if education_info else '',
            ]
            ws.append(row)

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Prepare response
        filename = f"enquiries_{start_date_str}_to_{end_date_str}.xlsx"
        response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    else:
        # Show form with today's date pre-filled in IST
        ist = pytz.timezone('Asia/Kolkata')
        today = timezone.now().astimezone(ist).date()

        return render(request, 'export_enquiries_form.html', {
            'today': today.strftime("%d-%m-%Y")
        })
        
# views.py
from django.views.generic import CreateView, UpdateView
from django.urls import reverse_lazy
from .models import MonthlyTarget
from .forms import MonthlyTargetForm


from django.utils.decorators import method_decorator


@method_decorator(manager_required, name='dispatch')
class MonthlyTargetCreateView(CreateView):
    model = MonthlyTarget
    form_class = MonthlyTargetForm
    template_name = 'monthly_target_form.html'
    success_url = reverse_lazy('enquiry:dashboard')  # or any success page

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Set Monthly Target'
        return context


@method_decorator(manager_required, name='dispatch')
class MonthlyTargetUpdateView(UpdateView):
    model = MonthlyTarget
    form_class = MonthlyTargetForm
    template_name = 'monthly_target_form.html'
    success_url = reverse_lazy('enquiry:dashboard')  # or any success page
    pk_url_kwarg = 'target_id'  # Optional custom ID name

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f"Edit Target for {self.object.get_month_display()} {self.object.year}"
        return context
    
# views.py
from django.http import JsonResponse
from django.utils.dateparse import parse_date
from django.shortcuts import render
from django.db.models import Q, Sum, F
from django.contrib.auth.decorators import login_required
from .models import Enquiry


@login_required
def due_fees_calendar_view(request):
    """
    Render the due fees calendar page with enhanced features
    """
    context = {
        'page_title': 'Due Fees Calendar',
        'page_description': 'Track student payment dues with interactive calendar view'
    }
    return render(request, 'due_fees_calendar.html', context)

from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.utils.dateparse import parse_date
from django.db.models import Q
from .models import Enquiry
from datetime import date
from django.http import JsonResponse
from django.utils.dateparse import parse_date
from django.contrib.auth.decorators import login_required
from datetime import date
from decimal import Decimal

from .models import Enquiry


@login_required
def due_fees_data_ajax(request):
    """
    AJAX endpoint to fetch students with due fees.
    Uses due_date if present, else falls back to enquiry_date.
    """
    try:
        date_str = request.GET.get('date')  # Expected in YYYY-MM-DD
        counsellor_filter = request.GET.get('counsellor')

        filter_date = parse_date(date_str) if date_str else None
        today = date.today()

        # Start with joined students who have outstanding balance
        students = Enquiry.objects.filter(
            status='joined',
            fees_balance__gt=0
        ).select_related('counsellor')

        # Optional: Filter by counsellor name
        if counsellor_filter:
            students = students.filter(counsellor__name=counsellor_filter)

        # Prepare final list after applying date logic
        filtered_students = []
        for student in students:
            due_date = student.due_date or student.enquiry_date

            # Skip if no date available
            if not due_date:
                continue

            # If date filter is applied, match it
            if filter_date and due_date != filter_date:
                continue

            # Determine status tag based on due_date
            if due_date < today:
                status_tag = 'overdue'
            elif due_date == today:
                status_tag = 'due_today'
            else:
                status_tag = 'upcoming'

            # Prepare safe data
            target_fees = float(student.target_fees) if student.target_fees else 0.0
            fees_due = float(student.fees_balance) if student.fees_balance else 0.0
            fees_paid = float(student.fees_paid) if student.fees_paid else 0.0

            filtered_students.append({
                'id': student.id,
                'name': student.name or 'N/A',
                'mobile': student.mobile or 'N/A',
                'counsellor': student.counsellor.name if student.counsellor else 'N/A',
                'counsellor_id': student.counsellor.id if student.counsellor else None,
                'target_fees': target_fees,
                'fees_paid': fees_paid,
                'fees_due': fees_due,
                'due_date': str(student.due_date) if student.due_date else '',
                'fallback_enquiry_date': str(student.enquiry_date) if student.enquiry_date else '',
                'status': status_tag,
                'course': dict(Enquiry.SUBJECT_CHOICES).get(student.subject, 'N/A'),
                'created_date': str(student.created_at)[:10] if student.created_at else '',
                'last_followup': str(student.followup_date) if student.followup_date else '',
            })

        return JsonResponse(filtered_students, safe=False)

    except Exception as e:
        return JsonResponse({
            'error': 'Failed to fetch due fees data',
            'details': str(e),
        }, status=500)




# views.py
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_protect
from django.contrib.auth.decorators import login_required
from .models import Enquiry, EducationInfo

@csrf_protect
@login_required
def update_education(request):
    if request.method == 'POST':
        try:
            enquiry_id = request.POST.get('enquiry_id')
            enquiry = Enquiry.objects.get(id=enquiry_id)
            
            # Get or create education info
            edu_info = enquiry.education_details.first() or EducationInfo(enquiry=enquiry)
            
            # Update fields from POST data
            edu_info.level = request.POST.get('education_level')
            edu_info.ug_degree = request.POST.get('ug_degree')
            edu_info.other_ug_degree_name = request.POST.get('other_ug_degree_name', '')
            edu_info.pg_degree = request.POST.get('pg_degree')
            edu_info.other_pg_degree_name = request.POST.get('other_pg_degree_name', '')
            edu_info.branch = request.POST.get('branch', '').upper()
            edu_info.year_of_passing = request.POST.get('year_of_passing')
            edu_info.percentage = request.POST.get('percentage')
            edu_info.college_name = request.POST.get('college_name', '')
            edu_info.college_place = request.POST.get('college_place', '')

            # Validate required fields
            if edu_info.level == 'ug' and not edu_info.ug_degree:
                return JsonResponse({
                    'status': 'error',
                    'message': 'UG degree is required for undergraduate level'
                }, status=400)
                
            if edu_info.level == 'pg' and not edu_info.pg_degree:
                return JsonResponse({
                    'status': 'error',
                    'message': 'PG degree is required for postgraduate level'
                }, status=400)

            # Additional validation for year and percentage
            try:
                year = int(edu_info.year_of_passing)
                current_year = 2024  # Should be dynamic in production
                if not (1900 <= year <= current_year + 1):
                    return JsonResponse({
                        'status': 'error',
                        'message': f'Year must be between 1900 and {current_year + 1}'
                    }, status=400)
            except ValueError:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Invalid year format'
                }, status=400)

            try:
                percentage = float(edu_info.percentage)
                if not (0 <= percentage <= 100):
                    return JsonResponse({
                        'status': 'error',
                        'message': 'Percentage must be between 0 and 100'
                    }, status=400)
            except ValueError:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Invalid percentage format'
                }, status=400)

            edu_info.save()
            
            return JsonResponse({
                'status': 'success',
                'message': 'Education details updated successfully'
            })
            
        except Enquiry.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': 'Enquiry not found'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=500)
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_protect
@login_required
def update_target_fees(request):
    if request.method == 'POST':
        try:
            enquiry_id = request.POST.get('enquiry_id')
            enquiry = Enquiry.objects.get(id=enquiry_id)
            
            # Check if status is 'joined' before allowing fee updates
            if enquiry.status != 'joined':
                return JsonResponse({
                    'status': 'error',
                    'message': 'Only joined enquiries can have target fees'
                }, status=400)
            
            target_fees = request.POST.get('target_fees')
            due_date = request.POST.get('due_date')
            
            # Validate required fields
            if not target_fees:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Target fees is required'
                }, status=400)
                
            if not due_date:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Due date is required'
                }, status=400)
            
            try:
                target_fees = float(target_fees)
                if target_fees <= 0:
                    return JsonResponse({
                        'status': 'error',
                        'message': 'Target fees must be a positive number'
                    }, status=400)
            except ValueError:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Invalid fee amount'
                }, status=400)
            
            # Update enquiry
            enquiry.target_fees = target_fees
            enquiry.due_date = due_date
            
            # Calculate balance if fees paid exists
            if enquiry.fees_paid:
                enquiry.fees_balance = max(0, target_fees - float(enquiry.fees_paid))
            
            enquiry.save()
            
            return JsonResponse({
                'status': 'success',
                'message': 'Target fees and due date updated successfully',
                'new_due_date': due_date,
                'fees_paid_percentage': enquiry.fees_paid_percentage if enquiry.target_fees else 0
            })
            
        except Enquiry.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': 'Enquiry not found'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=500)
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)






from django.views.generic import ListView
from .models import Enquiry

class InconsistentJoinedEnquiryListView(ListView):
    model = Enquiry
    template_name = 'inconsistent_joined_list.html'
    context_object_name = 'enquiries'

    def get_queryset(self):
        return Enquiry.objects.filter(status='joined', is_joined_batch=False)


from django.views.decorators.http import require_POST
from django.http import JsonResponse
from .models import Enquiry

@require_POST
def mark_joined_batch(request, enquiry_id):
    try:
        enquiry = Enquiry.objects.get(pk=enquiry_id, status='joined', is_joined_batch=False)
        enquiry.is_joined_batch = True
        enquiry.save()
        return JsonResponse({'success': True})
    except Enquiry.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Enquiry not found'})
    
from django.db.models import Count, Q, F, Sum, Case, When, Value, IntegerField
from django.db.models.functions import TruncDate
from django.shortcuts import render
from django.utils import timezone
from .models import Enquiry
from zoneinfo import ZoneInfo

def today_enquiries(request):
    india_tz = ZoneInfo("Asia/Kolkata")
    today = timezone.now().astimezone(india_tz).date()
    filter_by = request.GET.get('filter_by', 'created')

    # Base queryset for today
    if filter_by == 'created':
        enquiries = Enquiry.objects.annotate(
            local_created_date=TruncDate(F('created_at'), tzinfo=india_tz)
        ).filter(local_created_date=today)
    else:
        enquiries = Enquiry.objects.filter(enquiry_date=today)

    # Previous enquiries converted today
    previous_converted_today = Enquiry.objects.filter(
        enquiry_date__lt=today,
        joined_date=today,
        status='joined'
    )

    # All involved counsellor IDs
    all_counsellor_ids = set(
        enquiries.values_list('counsellor_id', flat=True)
    ).union(
        previous_converted_today.values_list('counsellor_id', flat=True)
    )

    # Breakdown per counsellor
    counsellor_breakdown = (
        enquiries
        .select_related('counsellor')
        .values('counsellor__id', 'counsellor__name')
        .annotate(
            total=Count('id'),

            # Walk-in Enquiries (only direct types)
            walkin=Count(Case(
                When(enquiry_type__in=['direct_walkin', 'someone_walkin'], then=Value(1)),
                output_field=IntegerField()
            )),
            walkin_joined=Count(Case(
                When(enquiry_type__in=['direct_walkin', 'someone_walkin'], status='joined', then=Value(1)),
                output_field=IntegerField()
            )),

            # Telephonic (including telephonic_to_walkin)
            telephonic=Count(Case(
                When(enquiry_type__in=['direct_telephonic', 'someone_telephonic', 'telephonic_to_walkin'], then=Value(1)),
                output_field=IntegerField()
            )),
            telephonic_joined=Count(Case(
                When(enquiry_type__in=['direct_telephonic', 'someone_telephonic', 'telephonic_to_walkin'], status='joined', then=Value(1)),
                output_field=IntegerField()
            )),

            # Other categories
            dropout=Count(Case(
                When(status='dropout', then=Value(1)),
                output_field=IntegerField()
            )),
            paid_enquiries=Count(Case(
                When(fees_paid__gt=0, then=Value(1)),
                output_field=IntegerField()
            )),
            fully_paid=Count(Case(
                When(target_fees__isnull=False, fees_paid__gte=F('target_fees'), then=Value(1)),
                output_field=IntegerField()
            )),
        )
        .order_by('counsellor__name')
    )

    # Previous conversions count map
    prev_conversion_map = previous_converted_today.values('counsellor_id').annotate(
        prev_converted=Count('id')
    )
    prev_map = {item['counsellor_id']: item['prev_converted'] for item in prev_conversion_map}

    # Append previous conversions to existing breakdown
    for block in counsellor_breakdown:
        cid = block['counsellor__id']
        block['previous_converted_today'] = prev_map.get(cid, 0)

    # Add counsellors who only had previous conversions
    existing_ids = {c['counsellor__id'] for c in counsellor_breakdown}
    only_conversion_ids = set(prev_map.keys()) - existing_ids

    if only_conversion_ids:
        extra_counsellors = (
            Enquiry.objects.filter(counsellor_id__in=only_conversion_ids)
            .values('counsellor__id', 'counsellor__name')
            .distinct()
        )

        for c in extra_counsellors:
            cid = c['counsellor__id']
            counsellor_breakdown = list(counsellor_breakdown)
            counsellor_breakdown.append({
                'counsellor__id': cid,
                'counsellor__name': c['counsellor__name'],
                'total': 0,
                'walkin': 0,
                'walkin_joined': 0,
                'telephonic': 0,
                'telephonic_joined': 0,
                'dropout': 0,
                'paid_enquiries': 0,
                'fully_paid': 0,
                'previous_converted_today': prev_map[cid],
            })

    context = {
        'enquiries': enquiries,
        'counsellor_breakdown': counsellor_breakdown,
        'filter_by': filter_by,
        'today': today,
    }

    return render(request, 'today_enquiries.html', context)





    
    
    
from django.shortcuts import render, get_object_or_404
from .models import Counsellor, Enquiry
from datetime import date
from django.contrib import messages

@manager_required
def previous_converted_list(request, counsellor_id):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    try:
        start_date = date.fromisoformat(start_date) if start_date else date.today()
        end_date = date.fromisoformat(end_date) if end_date else date.today()
    except ValueError:
        start_date = end_date = date.today()
        messages.error(request, "Invalid date format. Showing today's data.")

    if start_date > end_date:
        messages.error(request, "Start date must be before end date.")
        start_date, end_date = end_date, start_date  # Swap

    counsellor = get_object_or_404(Counsellor, id=counsellor_id)

    prev_converted_list = Enquiry.objects.filter(
        counsellor=counsellor,
        enquiry_date__lt=start_date,
        joined_date__range=(start_date, end_date),
        status='joined'
    )

    context = {
        'counsellor': counsellor,
        'prev_converted_list': prev_converted_list,
        'start_date': start_date,
        'end_date': end_date,
    }

    return render(request, 'previous_converted_list.html', context)



# views.py
from django.views import View
from django.shortcuts import render, get_object_or_404, redirect
from django.db import transaction
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.db.models import Q
from .models import Enquiry

def search_by_two_mobiles(request):
    q1 = request.GET.get('q1', '')
    q2 = request.GET.get('q2', '')
    results = []

    if q1 or q2:
        query1 = Enquiry.objects.filter(
            Q(mobile__icontains=q1) | Q(parent_number__icontains=q1)
        )
        query2 = Enquiry.objects.filter(
            Q(mobile__icontains=q2) | Q(parent_number__icontains=q2)
        )
        results = (query1 | query2).distinct().order_by('-created_at')

    return render(request, 'merge_two_mobiles.html', {
        'q1': q1,
        'q2': q2,
        'results': results
    })

class MergeEnquiriesView(View):
    def post(self, request, *args, **kwargs):
        selected_ids = request.POST.getlist('enquiry_ids')

        if len(selected_ids) != 2:
            messages.error(request, "Please select exactly two records to merge.")
            return redirect('enquiry:search_by_two_mobiles')

        primary_id, secondary_id = selected_ids
        primary = get_object_or_404(Enquiry, id=secondary_id)
        secondary = get_object_or_404(Enquiry, id=primary_id)

        if primary == secondary:
            messages.error(request, "Cannot merge a record with itself.")
            return redirect('enquiry:search_by_two_mobiles')

        allowed_types = ['someone_telephonic', 'someone_walkin']
        if secondary.enquiry_type not in allowed_types:
            messages.error(
                request,
                "The primary record must be of type 'Someone Telephonic' or 'Someone WalkIn'."
            )
            return redirect('enquiry:search_by_two_mobiles')

        try:
            with transaction.atomic():
                original_mobile = primary.mobile
                original_name = primary.name

                # Fields to merge from secondary to primary (excluding name and mobile)
                fields_to_merge = [
                    'subject', 'other_subject_name', 'status',
                    'enquiry_type', 'followup_date', 'enquiry_date', 'fees_paid_date',
                    'native_district_name', 'target_fees', 'fees_paid', 'due_date',
                    'is_joined_batch', 'joined_date', 'visit_type',
                    'parent_number', 'counsellor'
                ]

                for field in fields_to_merge:
                    secondary_value = getattr(secondary, field)
                    if secondary_value not in [None, '', 'None']:
                        setattr(primary, field, secondary_value)

                # Ensure name and mobile remain unchanged
                primary.name = original_name
                primary.mobile = original_mobile

                # Validate and save
                primary.full_clean()
                primary.save()

                # Delete secondary record
                secondary.delete()

                messages.success(request, "🎉 Enquiries merged successfully.")

        except ValidationError as e:
            messages.error(request, f"Validation error during merge: {e}")
        except Exception as e:
            messages.error(request, f"An error occurred: {e}")

        return redirect('enquiry:search_by_two_mobiles')
    




@manager_required
def enquiry_list_by_filter(request, counsellor_id, filter_type):
    from datetime import date
    from django.contrib import messages
    from django.shortcuts import get_object_or_404, render
    from enquiry.models import Enquiry, Counsellor

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    try:
        start_date = date.fromisoformat(start_date) if start_date else date.today()
        end_date = date.fromisoformat(end_date) if end_date else date.today()
    except ValueError:
        start_date = end_date = date.today()
        messages.error(request, "Invalid date format. Showing today's data.")

    if start_date > end_date:
        start_date, end_date = end_date, start_date
        messages.error(request, "Start date should be before end date.")

    counsellor = get_object_or_404(Counsellor, id=counsellor_id)

    # Consistent with summary view
    base_queryset = Enquiry.objects.filter(
        counsellor=counsellor,
        created_at__date__range=(start_date, end_date)
    )

    # Consistent filtering
    telephonic_types = ['direct_telephonic', 'someone_telephonic', 'telephonic_to_walkin']
    walkin_types = ['direct_walkin', 'someone_walkin']

    # Previous converted
    prev_converted_qs = Enquiry.objects.filter(
        counsellor=counsellor,
        enquiry_date__lt=start_date,
        joined_date__range=(start_date, end_date),
        status='joined'
    )

    # Filters map (aligned with your summary)
    filter_type_map = {
        'total': {
            'query': base_queryset,
            'label': 'Total Enquiries',
        },
        'telephonic': {
            'query': base_queryset.filter(enquiry_type__in=telephonic_types),
            'label': 'Telephonic Enquiries',
        },
        'telephonic_joined': {
            'query': base_queryset.filter(enquiry_type__in=telephonic_types, status='joined'),
            'label': 'Telephonic Joined',
        },
        'walkin': {
            'query': base_queryset.filter(enquiry_type__in=walkin_types),
            'label': 'Walk-in Enquiries',
        },
        'walkin_joined': {
            'query': base_queryset.filter(enquiry_type__in=walkin_types, status='joined'),
            'label': 'Walk-in Joined',
        },
        'joined': {
            'query': base_queryset.filter(status='joined'),
            'label': 'Joined Enquiries',
        },
        'pending': {
            'query': base_queryset.filter(status='pending'),
            'label': 'Pending Enquiries',
        },
        'dropout': {
            'query': base_queryset.filter(status='dropout'),
            'label': 'Dropout Enquiries',
        },
        'previous_converted': {
            'query': prev_converted_qs,
            'label': 'Previously Converted Enquiries',
        }
    }

    selected = filter_type_map.get(filter_type)

    if selected:
        enquiries = selected['query']
        label = selected['label']
    else:
        enquiries = Enquiry.objects.none()
        label = 'Unknown Filter'

    return render(request, 'enquiry_filter_list.html', {
        'enquiries': enquiries,
        'filter_type': filter_type,
        'label': label,
        'counsellor': counsellor,
        'start_date': start_date,
        'end_date': end_date,
    })
